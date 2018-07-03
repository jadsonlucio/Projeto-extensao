from numpy import NaN
from openpyxl import Workbook, load_workbook

from ..constantes import CAMINHO_METRICAS_DATABASE,CAMINHO_DATABASE,CAMINHO_SELECTED_DATABASE
from ..processamento import datas
from ..processamento import arquivo
from ..instancias.Instancias import get_instancia,add_instancia,instancias
from .exceptions.exception import tratamento_excessao,infoerroexception

array_tabelas=[]

class tabela():
    def __init__(self):
        self.arq_tabela = None
        self.nome_tabela = None

    # funcões de tabela

    def criar_tabela(self, nome_tabela):
        try:
            self.nome_tabela = nome_tabela
            self.arq_tabela = Workbook()
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def load_tabela(self, url, ready_only=False, data_only=False):
        try:
            self.arq_tabela = load_workbook(url, read_only=ready_only, data_only=data_only)
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def salvar_tabela(self, url=None):
        try:
            if(url==None):
                self.arq_tabela.save(CAMINHO_DATABASE + self.nome_tabela)
            else:
                self.arq_tabela.save(url)
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    # funcões de planilha

    def criar_planilha(self, index, nome_planilha):
        try:
            self.arq_tabela.create_sheet(nome_planilha, index)
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def get_column_planilha(self,index_planilha,index_column,min_row,max_row):
        try:
            resultado=self.copiar_planilha(index_planilha,min_col=index_column,
                                           max_col=index_column,min_row=min_row,max_row=max_row)
            return [valor[0] for valor in resultado]
        except Exception as e:
            tratamento_excessao("Erro")

    def get_row_planilha(self,index_planilha,index_row,min_col,max_col):
        try:
            resultado=self.copiar_planilha(index_planilha,min_row=index_row,
                                           max_row=index_row,min_col=min_col,max_col=max_col)
            return resultado[0]
        except Exception as e:
            tratamento_excessao("Erro")

    def copiar_planilha(self, index, min_row=1, max_row=None, min_col=1, max_col=None, mode='horizontal'):
        try:
            array = []
            planilha = self.arq_tabela.worksheets[index]
            if (max_row == None):
                max_row = planilha.max_row
            elif(max_row>planilha.max_row):
                max_row = planilha.max_row
            if (max_col == None):
                max_col = planilha.max_column
            elif(max_col>planilha.max_column):
                max_col = planilha.max_column
            if (mode == 'horizontal'):
                for row in planilha.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    linha = []
                    for cell in row:
                        linha.append(cell.value)
                    array.append(linha)
            if (mode == 'vertical'):
                for row in planilha.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    linha = []
                    for cell in row:
                        linha.append(cell.value)
                    array.append(linha)
            return array
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def salvar_planilha(self, index, array, row_ini=0, col_ini=0, NaN_fill=False):
        try:
            cont_rows = row_ini
            planilha = self.arq_tabela.worksheets[index]
            for row in array:
                cont_col = col_ini
                for cell in row:
                    if (NaN_fill == 0 and cell == NaN):
                        planilha.cell(row=cont_rows, column=cont_col, value=0)
                    elif (NaN_fill == -1 and cell == NaN):
                        new_cell = array[cont_col][cont_col - col_ini - 1]
                        planilha.cell(row=cont_rows, column=cont_col, value=new_cell)
                    else:
                        planilha.cell(row=cont_rows, column=cont_col, value=cell)

                    cont_col = cont_col + 1
                cont_rows = cont_rows + 1
        except Exception as e:
            tratamento_excessao(type_exception='Erro')


class database(tabela):

    # geters and seters

    def get_metricas(self):
        try:
            return self.metricas
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def set_metricas(self, type=None, **kwargs):
        try:
            if (type == 'file'):
                kwargs = arquivo.ler_arquivo_metricas(kwargs['url'])
            else:
                raise infoerroexception('opção invalida')
            self.nome_tabela = kwargs['nome_tabela']
            self.orientacao = kwargs['orientacao']
            self.periodo = kwargs['periodo']
            self.time_steps = int(kwargs['time_steps'])
            array_data = [int(valor) for valor in kwargs['data_inicial'].split('/')]
            self.data_inicial = datas.date(day=array_data[0], month=array_data[1], year=array_data[2])
            self.keys_array = [valor for valor in kwargs['keys_array'].split(',')]
            self.set_metricas_processamento()
        except infoerroexception as e:
            tratamento_excessao(type_exception='Info')
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def set_metricas_processamento(self):
        get_instancia('pre_processamento')._set_parameter(None, self, self.orientacao, self.periodo, self.time_steps, self.data_inicial,
                                self.keys_array)

    def __init__(self, orientacao_data, periodo, time_steps, data_inicial, keys_array):
        tabela.__init__(self)
        self.nome_tabela = 'database.xlsx'
        self.orientacao = orientacao_data
        self.periodo = periodo
        self.time_steps = time_steps
        self.data_inicial = data_inicial
        self.keys_array = keys_array

    #funções de operação com arquivos

    def verificar_arquivo(self):
        file=arquivo.abrir_arquivo(CAMINHO_SELECTED_DATABASE)
        array_file=arquivo.ler_array_arquivo(file)
        if(len(array_file)==1):
            return True
        else:
            return False

    def carregar_database(self):
        try:
            file=arquivo.abrir_arquivo(CAMINHO_SELECTED_DATABASE)
            array_file=arquivo.ler_array_arquivo(file)
            nome_planilha=array_file[0]

            planilha_url=CAMINHO_DATABASE+"//"+nome_planilha
            metricas_url=CAMINHO_METRICAS_DATABASE+"//"+nome_planilha+".txt"

            self.load_tabela(planilha_url)
            self.set_metricas(type="file",url=metricas_url)

            if("frame_opcoes" in instancias.keys()):
                frame_opcoes=get_instancia("frame_opcoes")
                frame_opcoes.set_value_metricabox()
        except Exception as e:
            tratamento_excessao(type_exception='Erro')

    def adcionar_arquivos(self, index, caminhos_arquivos):
        try:
            self.arquivos_list = datas.organizar_arquivos(caminhos_arquivos)
            if (len(self.arquivos_list) == 0):
                raise infoerroexception("Nenhum arquivo valido selecionado")
            if (self.arq_tabela == None):
                raise infoerroexception("Nenhum arquivo para gravacao encontrado")
            for arquivo_list in self.arquivos_list:
                data_ini = datas.date(year=datas.data_inicial[0], month=datas.data_inicial[1], day=1)
                data_fim = datas.date(year=arquivo_list[2], month=arquivo_list[3], day=1)
                row_ini = datas.operacao_datas('-', data_ini=data_ini, data_fim=data_fim) * 144 + 1
                tab = tabela()
                tab.load_tabela(arquivo_list[1])
                array = tab.copiar_planilha(0, min_row=6, min_col=3)
                self.salvar_planilha(index, array, row_ini, col_ini=1, NaN_fill=0)
            return 1
        except infoerroexception as e:
            tratamento_excessao(type_exception='Info')
        except Exception as e:
            tratamento_excessao(type_exception='Erro')
            return 0

def criar_tabela():
    tab=tabela()
    array_tabelas.append(tab)
    return tab

def criar_instancia_database():
    inst_database=database('orizontal', 'minuto', 10, datas.date(day=1, month=1, year=2013), [])
    add_instancia("database",inst_database)